"""
Admin dashboard page
"""
import streamlit as st
from datetime import datetime
import os

from frontend.components.styles import get_main_app_styles
from frontend.components.utils import format_time_12h
from config import Config
from backend.database import db_queries
from backend.utils.error_handler import log_user_action, check_session_timeout
from backend.services.case_service import case_service
from backend.services.validation_service import validation_service

def render_admin_page():
    """Main admin page renderer"""
    # Apply styles
    st.markdown(get_main_app_styles(), unsafe_allow_html=True)
    
    # Check session timeout
    try:
        if not check_session_timeout():
            st.error("⏰ Your session has expired. Please login again.")
            st.session_state.clear()
            st.rerun()
    except:
        pass  # If check_session_timeout not available
    
    # Render sidebar
    render_admin_sidebar()
    
    # Main content - Professional header
    st.markdown("""
    <div style='background: linear-gradient(135deg, #1565c0 0%, #667eea 100%); 
                padding: 2rem; border-radius: 12px; margin-bottom: 2rem; 
                box-shadow: 0 4px 16px rgba(0,0,0,0.1);'>
        <h1 style='color: white; margin: 0; font-size: 2.5rem;'>🏢 Admin Dashboard</h1>
        <p style='color: rgba(255,255,255,0.9); margin: 0.5rem 0 0 0; font-size: 1.1rem;'>
            AI Missing Person Identification System
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    # User info display - Professional card
    user_info = {"name": "AI MISSING PERSON IDENTIFICATION SYSTEM", "area": "MIT", "city": "Chh. Sambhajinagar"}
    role = "Admin"
    role_colour = "#e74c3c"
    role_badge = (
        f'<span style="background:{role_colour}; color:white; padding:6px 14px; '
        f'border-radius:20px; font-size:12px; font-weight:600; display:inline-block;">{role}</span>'
    )
    st.markdown(f"""
    <div style='background: white; padding: 1.5rem; border-radius: 8px; 
                box-shadow: 0 2px 8px rgba(0,0,0,0.08); margin-bottom: 2rem;'>
        <div style='display: flex; justify-content: space-between; align-items: center;'>
            <div>
                <p style='color: #546e7a; margin: 0; font-size: 0.9rem;'>📍 Location</p>
                <p style='color: #1a237e; margin: 0.5rem 0 0 0; font-size: 1.1rem; font-weight: 600;'>
                    {user_info["area"]}, {user_info["city"]}
                </p>
            </div>
            <div>{role_badge}</div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # Dashboard options - Professional layout
    st.markdown("### 📊 Dashboard Functions")
    dashboard_option = st.selectbox("Select Function:", [
        "Register New Case",
        "All Cases", 
        "Match Cases",
        "Send Message",
        "Map",
        "User Management",
        "Email Configuration"
    ], label_visibility="collapsed")
    
    # Route to appropriate function
    if dashboard_option == "Register New Case":
        render_register_case()
    elif dashboard_option == "All Cases":
        render_all_cases()
    elif dashboard_option == "Match Cases":
        render_match_cases()
    elif dashboard_option == "Send Message":
        render_send_message()
    elif dashboard_option == "Map":
        render_map()
    elif dashboard_option == "User Management":
        render_user_management()
    elif dashboard_option == "Email Configuration":
        render_email_configuration()


def render_admin_sidebar():
    """Render admin sidebar"""
    with st.sidebar:
        # Professional header
        st.markdown("""
        <div style='background: linear-gradient(135deg, #1565c0 0%, #667eea 100%); 
                    padding: 1.5rem; border-radius: 8px; margin-bottom: 1.5rem;
                    box-shadow: 0 2px 8px rgba(0,0,0,0.1);'>
            <h3 style='color: white; margin: 0; font-size: 1.3rem;'>👋 Welcome, Officer</h3>
            <p style='color: rgba(255,255,255,0.9); margin: 0.5rem 0 0 0; font-size: 0.9rem;'>
                Admin Access
            </p>
        </div>
        """, unsafe_allow_html=True)
        
        # User information - Clean layout
        st.markdown("#### 👤 User Information")
        st.markdown(f"""
        <div style='background: #f5f7fa; padding: 1rem; border-radius: 8px; margin-bottom: 1rem;'>
            <p style='margin: 0.5rem 0; font-size: 0.9rem;'><strong>Role:</strong> Admin</p>
            <p style='margin: 0.5rem 0; font-size: 0.9rem;'><strong>Username:</strong> {st.session_state.get('username', 'N/A')}</p>
            <p style='margin: 0.5rem 0; font-size: 0.9rem;'><strong>Status:</strong> ✅ Authenticated</p>
            <p style='margin: 0.5rem 0; font-size: 0.9rem;'><strong>Version:</strong> {Config.APP_VERSION}</p>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("---")
        
        # Logout button
        if st.button("🚪 Logout", use_container_width=True, type="secondary"):
            try:
                log_user_action("Logout", st.session_state.get("username"))
            except:
                pass
            st.session_state["authentication_status"] = False
            st.session_state["user_role"] = None
            st.session_state.clear()
            st.rerun()


def render_register_case():
    """Render register new case form"""
    st.markdown("---")
    st.markdown("""
    <div style='background: linear-gradient(135deg, #27ae60 0%, #229954 100%); 
                padding: 1.5rem; border-radius: 8px; margin-bottom: 1.5rem;
                box-shadow: 0 2px 8px rgba(0,0,0,0.1);'>
        <h2 style='color: white; margin: 0; font-size: 1.8rem;'>📝 Register New Missing Person Case</h2>
        <p style='color: rgba(255,255,255,0.9); margin: 0.5rem 0 0 0;'>
            Register a new missing person case with photo and details
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    st.info("📋 This form allows you to register a new missing person case with photo and details.")

    image_col, form_col = st.columns([1, 1.2], gap="large")
    face_mesh = None
    face_detected = False
    unique_id = None

    with image_col:
        import uuid
        from backend.utils.image_utils import image_obj_to_numpy, extract_face_mesh_landmarks
        
        st.markdown("#### 📸 Upload Photo")
        image_obj = st.file_uploader(
            "Upload Photo", type=["jpg", "jpeg", "png"], key="register_case_img", label_visibility="collapsed"
        )
        if image_obj:
            unique_id = str(uuid.uuid4())
            
            with st.spinner("Processing image..."):
                os.makedirs("./resources", exist_ok=True)
                uploaded_file_path = "./resources/" + unique_id + ".jpg"
                with open(uploaded_file_path, "wb") as f:
                    f.write(image_obj.read())
                
                image_obj.seek(0)
                st.image(image_obj, width=300, caption="Uploaded Photo")
                image_obj.seek(0)
                
                image_np = image_obj_to_numpy(image_obj)
                landmarks = extract_face_mesh_landmarks(image_np)

                if landmarks is not None:
                    face_mesh = landmarks
                    face_detected = True
                    st.success("✅ Face detected successfully!")
                else:
                    st.warning("⚠️ No face detected. Please upload a clear photo with a visible face.")

    with form_col:
        if face_detected:
            import json
            from backend.models.data_models import RegisteredCases
            from frontend.components.utils import get_indian_states_and_cities
            
            st.markdown("#### 📋 Case Details")
            
            # Personal Information
            st.markdown("**👤 Personal Information**")
            col1, col2 = st.columns(2)
            with col1:
                name = st.text_input("Person's Name *", key="reg_name")
            with col2:
                father_name = st.text_input("Father's Name", key="reg_father")
            age = st.text_input("Age", key="reg_age")
            
            # Complainant Information
            st.markdown("**📞 Complainant Information**")
            col1, col2 = st.columns(2)
            with col1:
                complainant_name = st.text_input("Your Name *", key="reg_comp_name")
            with col2:
                complainant_mobile = st.text_input("Mobile Number *", key="reg_comp_mobile", max_chars=10)
            complainant_email = st.text_input("Email", key="reg_comp_email")
            
            # Location Information
            st.markdown("**📍 Location Information**")
            indian_states, state_cities = get_indian_states_and_cities()
            
            col1, col2 = st.columns(2)
            with col1:
                state = st.selectbox("State *", [""] + sorted(indian_states), key="reg_state")
            with col2:
                if state and state in state_cities:
                    city = st.selectbox("City *", [""] + sorted(state_cities[state]), key="reg_city")
                else:
                    city = st.text_input("City *", key="reg_city_text")
            
            last_seen = st.text_input("Last Seen Date/Location *", key="reg_last_seen")
            
            # Additional Information
            st.markdown("**📝 Additional Information**")
            birth_marks = st.text_area("Identifying Marks/Birth Marks", key="reg_marks", height=80)
            
            if st.button("💾 Register Case", type="primary", use_container_width=True):
                # Validation using service
                is_valid, error_msg = validation_service.validate_case_registration(
                    name, complainant_name, complainant_mobile, city, last_seen
                )
                
                if not is_valid:
                    st.error(error_msg)
                else:
                    try:
                        # Create case
                        case = RegisteredCases(
                            id=unique_id,
                            submitted_by=st.session_state.get("username", "admin"),
                            name=name,
                            father_name=father_name or "",
                            age=age or "",
                            complainant_name=complainant_name,
                            complainant_mobile=complainant_mobile,
                            complainant_email=complainant_email or "",
                            adhaar_card="",
                            last_seen=last_seen,
                            address="",
                            state=state or "",
                            city=city,
                            face_mesh=json.dumps(face_mesh),
                            status="NF",
                            birth_marks=birth_marks or "",
                            matched_with=None
                        )
                        
                        db_queries.register_new_case(case)
                        st.success("✅ Case registered successfully!")
                        st.balloons()
                        
                        try:
                            log_user_action("Case registered", st.session_state.get("username"))
                        except:
                            pass
                        
                        # Clear form
                        st.info("🔄 Refresh the page to register another case.")
                        
                    except Exception as e:
                        st.error(f"❌ Error registering case: {str(e)}")


def render_all_cases():
    """Render all cases view - separated into Admin and Public sections"""
    st.markdown("---")
    st.markdown("""
    <div style='background: linear-gradient(135deg, #1565c0 0%, #667eea 100%); 
                padding: 1.5rem; border-radius: 8px; margin-bottom: 1.5rem;
                box-shadow: 0 2px 8px rgba(0,0,0,0.1);'>
        <h2 style='color: white; margin: 0; font-size: 1.8rem;'>📋 All Cases</h2>
        <p style='color: rgba(255,255,255,0.9); margin: 0.5rem 0 0 0;'>
            View and manage all registered cases and public submissions
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    try:
        # Filter options - Professional layout
        st.markdown("### 🔍 Filter & Search")
        col1, col2 = st.columns(2)
        
        with col1:
            status_filter = st.selectbox(
                "Filter by Status",
                ["All", "Not Found", "Found"],
                key="all_cases_status"
            )
        
        with col2:
            search_term = st.text_input("Search by name or location", "", key="all_cases_search")
        
        st.markdown("---")
        
        # Get all cases (both registered and public)
        all_reg = db_queries.get_all_registered_cases()
        all_pub = db_queries.fetch_public_cases(train_data=False, status="All")
        
        # Apply status filter using service
        registered_cases = case_service.filter_cases_by_status(all_reg, status_filter)
        public_submissions = case_service.filter_cases_by_status(all_pub, status_filter)
        
        # Apply search filter using service
        registered_cases = case_service.filter_registered_cases_by_search(registered_cases, search_term)
        public_submissions = case_service.filter_public_cases_by_search(public_submissions, search_term)
        
        # Sort by submission date using service
        registered_cases = case_service.sort_cases_by_date(registered_cases)
        public_submissions = case_service.sort_public_cases_by_date(public_submissions)
        
        # Calculate statistics using service
        stats = case_service.calculate_statistics(registered_cases, public_submissions)
        
        # Display statistics - Professional cards
        st.markdown("### 📊 Statistics")
        col1, col2, col3, col4, col5 = st.columns(5)
        
        with col1:
            st.markdown(f"""
            <div style='background: linear-gradient(135deg, #3498db 0%, #2980b9 100%); 
                        padding: 1.5rem; border-radius: 8px; text-align: center;
                        box-shadow: 0 2px 8px rgba(0,0,0,0.1);'>
                <p style='color: rgba(255,255,255,0.8); margin: 0; font-size: 0.9rem;'>Total Cases</p>
                <p style='color: white; margin: 0.5rem 0 0 0; font-size: 2rem; font-weight: bold;'>
                    {stats['total_cases']}
                </p>
            </div>
            """, unsafe_allow_html=True)
        
        with col2:
            st.markdown(f"""
            <div style='background: linear-gradient(135deg, #27ae60 0%, #229954 100%); 
                        padding: 1.5rem; border-radius: 8px; text-align: center;
                        box-shadow: 0 2px 8px rgba(0,0,0,0.1);'>
                <p style='color: rgba(255,255,255,0.8); margin: 0; font-size: 0.9rem;'>Found</p>
                <p style='color: white; margin: 0.5rem 0 0 0; font-size: 2rem; font-weight: bold;'>
                    {stats['found_count']}
                </p>
            </div>
            """, unsafe_allow_html=True)
        
        with col3:
            st.markdown(f"""
            <div style='background: linear-gradient(135deg, #e74c3c 0%, #c0392b 100%); 
                        padding: 1.5rem; border-radius: 8px; text-align: center;
                        box-shadow: 0 2px 8px rgba(0,0,0,0.1);'>
                <p style='color: rgba(255,255,255,0.8); margin: 0; font-size: 0.9rem;'>Not Found</p>
                <p style='color: white; margin: 0.5rem 0 0 0; font-size: 2rem; font-weight: bold;'>
                    {stats['not_found_count']}
                </p>
            </div>
            """, unsafe_allow_html=True)
        
        with col4:
            st.markdown(f"""
            <div style='background: linear-gradient(135deg, #f39c12 0%, #d68910 100%); 
                        padding: 1.5rem; border-radius: 8px; text-align: center;
                        box-shadow: 0 2px 8px rgba(0,0,0,0.1);'>
                <p style='color: rgba(255,255,255,0.8); margin: 0; font-size: 0.9rem;'>Admin</p>
                <p style='color: white; margin: 0.5rem 0 0 0; font-size: 2rem; font-weight: bold;'>
                    {stats['admin_count']}
                </p>
            </div>
            """, unsafe_allow_html=True)
        
        with col5:
            st.markdown(f"""
            <div style='background: linear-gradient(135deg, #9b59b6 0%, #8e44ad 100%); 
                        padding: 1.5rem; border-radius: 8px; text-align: center;
                        box-shadow: 0 2px 8px rgba(0,0,0,0.1);'>
                <p style='color: rgba(255,255,255,0.8); margin: 0; font-size: 0.9rem;'>Public</p>
                <p style='color: white; margin: 0.5rem 0 0 0; font-size: 2rem; font-weight: bold;'>
                    {stats['public_count']}
                </p>
            </div>
            """, unsafe_allow_html=True)
        
        st.markdown("---")
        
        # ========================================
        # SECTION 1: ADMIN REGISTERED CASES
        # ========================================
        st.markdown(f"### 🏢 Admin Registered Cases ({len(registered_cases)})")
        
        if registered_cases:
            for idx, data in enumerate(registered_cases):
                try:
                    case_id = data.id
                    name = data.name
                    location = data.city or 'Unknown'
                    status = data.status
                    
                    status_emoji, status_text = case_service.get_status_display(status)
                    
                    # Create expander title
                    title = f"{status_emoji} {name} - {location} ({status_text})"
                    
                    with st.expander(title, expanded=False):
                        col_photo, col_details, col_actions = st.columns([1, 2, 1])
                        
                        with col_photo:
                            try:
                                st.image(f"./resources/{case_id}.jpg", width=150)
                            except:
                                st.warning("📷 No image")
                        
                        with col_details:
                            st.markdown("**🏢 Registered Case**")
                            st.write(f"**Name:** {data.name}")
                            st.write(f"**Age:** {data.age or 'Unknown'}")
                            st.write(f"**Last Seen:** {data.last_seen}")
                            st.write(f"**City:** {data.city or 'Not provided'}")
                            st.write(f"**State:** {data.state or 'Not provided'}")
                            
                            st.markdown("**👤 Complainant**")
                            st.write(f"**Name:** {data.complainant_name}")
                            st.write(f"**Mobile:** {data.complainant_mobile}")
                            if data.complainant_email:
                                st.write(f"**Email:** {data.complainant_email}")
                            
                            if data.description:
                                st.markdown("**📝 Description**")
                                st.write(data.description)
                            
                            if data.birth_marks:
                                st.markdown("**🔍 Identifying Marks**")
                                st.write(data.birth_marks)
                            
                            st.markdown("**📊 Status**")
                            st.write(f"**Status:** {status_text}")
                            st.write(f"**Submitted On:** {format_time_12h(data.submitted_on)}")
                        
                        with col_actions:
                            st.markdown("**⚙️ Actions:**")
                            
                            # Edit button
                            if st.button("✏️ Edit", key=f"edit_admin_{case_id}_{idx}"):
                                st.session_state[f"editing_{case_id}"] = True
                                st.rerun()
                            
                            if status == "NF":
                                if st.button("✅ Mark Found", key=f"found_admin_{case_id}_{idx}"):
                                    db_queries.update_registered_case(case_id, {"status": "F"})
                                    st.success("Marked as Found!")
                                    st.rerun()
                            else:
                                if st.button("❌ Mark Not Found", key=f"nf_admin_{case_id}_{idx}"):
                                    db_queries.update_registered_case(case_id, {"status": "NF"})
                                    st.success("Marked as Not Found!")
                                    st.rerun()
                            
                            if st.button("🗑️ Delete", key=f"del_admin_{case_id}_{idx}"):
                                try:
                                    db_queries.delete_registered_case(case_id)
                                    st.success("Deleted!")
                                    st.rerun()
                                except Exception as e:
                                    st.error(f"Error: {str(e)}")
                        
                        # Edit form modal
                        if st.session_state.get(f"editing_{case_id}", False):
                            st.markdown("---")
                            st.markdown("### ✏️ Edit Case Details")
                            
                            from frontend.components.utils import get_indian_states_and_cities
                            
                            # Create edit form
                            with st.form(key=f"edit_form_{case_id}"):
                                st.markdown("**Personal Information**")
                                edit_name = st.text_input("Person's Name *", value=data.name, key=f"edit_name_{case_id}")
                                edit_father_name = st.text_input("Father's Name", value=data.father_name or "", key=f"edit_father_{case_id}")
                                edit_age = st.text_input("Age", value=data.age or "", key=f"edit_age_{case_id}")
                                
                                st.markdown("**Complainant Information**")
                                edit_comp_name = st.text_input("Complainant Name *", value=data.complainant_name, key=f"edit_comp_name_{case_id}")
                                edit_comp_mobile = st.text_input("Complainant Mobile *", value=data.complainant_mobile, max_chars=10, key=f"edit_comp_mobile_{case_id}")
                                edit_comp_email = st.text_input("Complainant Email", value=data.complainant_email or "", key=f"edit_comp_email_{case_id}")
                                
                                st.markdown("**Location Information**")
                                indian_states, state_cities = get_indian_states_and_cities()
                                
                                current_state = data.state or ""
                                state_index = 0
                                if current_state and current_state in indian_states:
                                    state_index = sorted(indian_states).index(current_state) + 1
                                
                                edit_state = st.selectbox("State *", [""] + sorted(indian_states), index=state_index, key=f"edit_state_{case_id}")
                                
                                if edit_state and edit_state in state_cities:
                                    current_city = data.city or ""
                                    city_index = 0
                                    if current_city and current_city in state_cities[edit_state]:
                                        city_index = sorted(state_cities[edit_state]).index(current_city) + 1
                                    edit_city = st.selectbox("City *", [""] + sorted(state_cities[edit_state]), index=city_index, key=f"edit_city_{case_id}")
                                else:
                                    edit_city = st.text_input("City *", value=data.city or "", key=f"edit_city_text_{case_id}")
                                
                                edit_last_seen = st.text_input("Last Seen Date/Location *", value=data.last_seen, key=f"edit_last_seen_{case_id}")
                                
                                st.markdown("**Additional Information**")
                                edit_birth_marks = st.text_area("Identifying Marks/Birth Marks", value=data.birth_marks or "", key=f"edit_marks_{case_id}")
                                
                                col_save, col_cancel = st.columns(2)
                                
                                with col_save:
                                    submit_edit = st.form_submit_button("💾 Save Changes", type="primary")
                                
                                with col_cancel:
                                    cancel_edit = st.form_submit_button("❌ Cancel")
                                
                                if submit_edit:
                                    # Validation using service
                                    is_valid, error_msg = validation_service.validate_case_update(
                                        edit_name, edit_comp_name, edit_comp_mobile, edit_city, edit_last_seen
                                    )
                                    
                                    if not is_valid:
                                        st.error(error_msg)
                                    else:
                                        try:
                                            # Update case
                                            update_fields = {
                                                "name": edit_name,
                                                "father_name": edit_father_name,
                                                "age": edit_age,
                                                "complainant_name": edit_comp_name,
                                                "complainant_mobile": edit_comp_mobile,
                                                "complainant_email": edit_comp_email,
                                                "state": edit_state,
                                                "city": edit_city,
                                                "last_seen": edit_last_seen,
                                                "birth_marks": edit_birth_marks
                                            }
                                            
                                            db_queries.update_registered_case(case_id, update_fields)
                                            st.success("✅ Case updated successfully!")
                                            
                                            try:
                                                log_user_action(f"Case edited: {case_id}", st.session_state.get("username"))
                                            except:
                                                pass
                                            
                                            # Clear editing state
                                            del st.session_state[f"editing_{case_id}"]
                                            st.rerun()
                                        except Exception as e:
                                            st.error(f"❌ Error updating case: {str(e)}")
                                
                                if cancel_edit:
                                    # Clear editing state
                                    del st.session_state[f"editing_{case_id}"]
                                    st.rerun()
                except Exception as e:
                    st.error(f"❌ Error displaying case {idx}: {str(e)}")
        else:
            st.info("No admin registered cases match the filter.")
        
        st.markdown("---")
        
        # ========================================
        # SECTION 2: PUBLIC SUBMISSIONS
        # ========================================
        st.markdown(f"### 👥 Public Submission Cases ({len(public_submissions)})")
        
        if public_submissions:
            for idx, data in enumerate(public_submissions):
                try:
                    case_id = data.id
                    name = data.name if (hasattr(data, 'name') and data.name) else 'Unknown'
                    location = data.location
                    status = data.status
                    
                    status_emoji, status_text = case_service.get_status_display(status)
                    
                    # Create expander title
                    title = f"{status_emoji} {name} - {location} ({status_text})"
                    
                    with st.expander(title, expanded=False):
                        col_photo, col_details, col_actions = st.columns([1, 2, 1])
                        
                        with col_photo:
                            try:
                                st.image(f"./resources/{case_id}.jpg", width=150)
                            except:
                                st.warning("📷 No image")
                        
                        with col_details:
                            st.markdown("**👥 Public Submission**")
                            if hasattr(data, 'name') and data.name:
                                st.write(f"**Person Name:** {data.name}")
                            st.write(f"**Location Sighted:** {data.location}")
                            if hasattr(data, 'city') and data.city:
                                st.write(f"**City:** {data.city}")
                            if hasattr(data, 'state') and data.state:
                                st.write(f"**State:** {data.state}")
                            
                            st.markdown("**👤 Submitted By**")
                            st.write(f"**Name:** {data.submitted_by or 'Anonymous'}")
                            st.write(f"**Mobile:** {data.mobile}")
                            if hasattr(data, 'email') and data.email:
                                st.write(f"**Email:** {data.email}")
                            
                            if hasattr(data, 'submitted_on') and data.submitted_on:
                                st.write(f"**Submitted On:** {format_time_12h(data.submitted_on)}")
                            
                            if hasattr(data, 'birth_marks') and data.birth_marks:
                                st.markdown("**📝 Description**")
                                st.write(data.birth_marks)
                            
                            st.markdown("**📊 Status**")
                            st.write(f"**Status:** {status_text}")
                        
                        with col_actions:
                            st.markdown("**⚙️ Actions:**")
                            if status == "NF":
                                if st.button("✅ Mark Found", key=f"found_public_{case_id}_{idx}"):
                                    db_queries.update_public_submission(case_id, {"status": "F"})
                                    st.success("Marked as Found!")
                                    st.rerun()
                            else:
                                if st.button("❌ Mark Not Found", key=f"nf_public_{case_id}_{idx}"):
                                    db_queries.update_public_submission(case_id, {"status": "NF"})
                                    st.success("Marked as Not Found!")
                                    st.rerun()
                            
                            if st.button("🗑️ Delete", key=f"del_public_{case_id}_{idx}"):
                                try:
                                    db_queries.delete_public_submission(case_id)
                                    st.success("Deleted!")
                                    st.rerun()
                                except Exception as e:
                                    st.error(f"Error: {str(e)}")
                except Exception as e:
                    st.error(f"❌ Error displaying submission {idx}: {str(e)}")
        else:
            st.info("No public submissions match the filter.")
    
    except Exception as e:
        st.error(f"❌ Error loading cases: {str(e)}")
        import traceback
        st.code(traceback.format_exc())


def render_match_cases():
    """Render match cases interface"""
    st.markdown("---")
    st.subheader("🎯 AI-Powered Face Matching")
    st.info("Match registered cases with public submissions using AI face recognition.")
    
    from backend.services import train_service, match_service
    
    # Set default threshold
    distance_threshold = 3.0
    
    st.markdown("---")
    
    # Initialize session state for match results
    if "match_results" not in st.session_state:
        st.session_state.match_results = None
    
    if st.button("🔄 Run Face Matching", type="primary"):
        with st.spinner("🤖 Training AI model..."):
            try:
                # Train model
                result = train_service.train(st.session_state.get("username", "admin"))
                
                if result["status"]:
                    st.success(f"✅ {result['message']}")
                else:
                    st.error(f"❌ Training failed: {result['message']}")
                    st.stop()
                
                # Run matching
                st.write("### 🔍 Finding Matches...")
                matched_result = match_service.match(distance_threshold=distance_threshold)
                
                # Store results in session state
                st.session_state.match_results = matched_result
            except Exception as e:
                st.error(f"❌ Error during matching: {str(e)}")
                import traceback
                st.code(traceback.format_exc())
    
    # Display results after spinner completes (OUTSIDE button block)
    if st.session_state.match_results:
        matched_result = st.session_state.match_results
        
        if matched_result["status"]:
            match_details = matched_result.get("details", [])
            summary = matched_result.get("summary", {})
            
            if not match_details:
                st.info("🤷‍♂️ No matches found.")
                st.write("**Try:**")
                st.write("- Adding more cases")
                st.write("- Ensuring photos are clear")
                
                # Show statistics
                st.markdown("---")
                st.markdown("### 📊 Statistics")
                col1, col2 = st.columns(2)
                with col1:
                    st.metric("Registered Cases", summary.get('registered_cases_count', 0))
                with col2:
                    st.metric("Public Submissions", summary.get('public_submissions_count', 0))
            else:
                total_matches = summary.get("total_matches", len(match_details))
                st.success(f"🎉 Found {total_matches} potential matches!")
                
                st.write(f"- Registered Cases: {summary.get('registered_cases_count', 0)}")
                st.write(f"- Public Submissions: {summary.get('public_submissions_count', 0)}")
                
                st.markdown("---")
                st.markdown("### 🎯 Match Results")
                
                for i, match in enumerate(match_details):
                    reg_id = match["registered_id"]
                    pub_id = match["public_id"]
                    confidence = match["confidence"]
                    distance = match["distance"]
                    reg_details = match["registered_details"]
                    pub_details = match["public_details"]
                    
                    st.markdown(f"#### Match {i+1}: {reg_details.get('name', 'Unknown')} ↔ {pub_details.get('location', 'Unknown')} ({confidence:.1f}% confidence)")
                    
                    # Get full registered case details
                    reg_case_full = db_queries.get_registered_case_detail(reg_id)
                    
                    st.markdown("**🏢 Registered Case Information**")
                    
                    # Display registered case with photo
                    col_photo, col_info = st.columns([1, 2])
                    
                    with col_photo:
                        try:
                            st.image(f"./resources/{reg_id}.jpg", width=200)
                        except:
                            st.warning("📷 No image")
                    
                    with col_info:
                        if reg_case_full:
                            reg_name, reg_mobile, reg_email, reg_age, reg_last_seen, reg_marks, reg_comp_name, reg_city, reg_state, reg_desc = reg_case_full[0]
                            
                            st.write(f"**Missing Person Name:** {reg_name}")
                            st.write(f"**Age:** {reg_age or 'Unknown'}")
                            st.write(f"**Last Seen:** {reg_last_seen}")
                            st.write(f"**City:** {reg_city or 'Not provided'}")
                            st.write(f"**State:** {reg_state or 'Not provided'}")
                            
                            st.markdown("**Complainant Details:**")
                            st.write(f"**Name:** {reg_comp_name}")
                            st.write(f"**Mobile:** {reg_mobile}")
                            if reg_email:
                                st.write(f"**Email:** {reg_email}")
                            
                            if reg_marks:
                                st.write(f"**Identifying Marks:** {reg_marks}")
                        else:
                            st.write(f"**Name:** {reg_details.get('name', 'Unknown')}")
                            st.write(f"**Last Seen:** {reg_details.get('last_seen', 'Unknown')}")
                    
                    st.markdown("---")
                    st.markdown("**👥 Public Submission Information**")
                    
                    # Display public submission with photo
                    col_photo_pub, col_info_pub = st.columns([1, 2])
                    
                    with col_photo_pub:
                        try:
                            st.image(f"./resources/{pub_id}.jpg", width=200)
                        except:
                            st.warning("📷 No image")
                    
                    with col_info_pub:
                        pub_case_full = db_queries.get_public_case_detail(pub_id)
                        
                        if pub_case_full:
                            pub_name, pub_location, pub_submitted_by, pub_mobile, pub_email, pub_marks, pub_city, pub_state = pub_case_full[0]
                            
                            st.write(f"**Location Sighted:** {pub_location}")
                            if pub_city:
                                st.write(f"**City:** {pub_city}")
                            if pub_state:
                                st.write(f"**State:** {pub_state}")
                            
                            st.markdown("**Submitted By:**")
                            st.write(f"**Name:** {pub_submitted_by or 'Anonymous'}")
                            st.write(f"**Mobile:** {pub_mobile}")
                            if pub_email:
                                st.write(f"**Email:** {pub_email}")
                            
                            if pub_marks:
                                st.write(f"**Description:** {pub_marks}")
                        else:
                            st.write(f"**Location:** {pub_details.get('location', 'Unknown')}")
                            st.write(f"**Submitted By:** {pub_details.get('submitted_by', 'Unknown')}")
                    
                    st.markdown("---")
                    st.markdown("**📊 Match Details**")
                    col_conf, col_dist = st.columns(2)
                    
                    with col_conf:
                        st.metric("Confidence", f"{confidence:.1f}%")
                        st.progress(confidence / 100)
                    
                    with col_dist:
                        st.metric("Distance", f"{distance:.3f}")
                    
                    st.markdown("---")
                    
                    # Confirm match button
                    col_confirm = st.columns(1)[0]
                    with col_confirm:
                        confirm_clicked = st.button(f"✅ Confirm Match", key=f"confirm_{reg_id}_{pub_id}", type="primary")
                    
                    if confirm_clicked:
                        try:
                            # Update status to Found for both cases
                            db_queries.update_found_status(reg_id, pub_id)
                            st.success("🎉 Match confirmed! Both records marked as 'Found'.")
                            st.balloons()
                            
                            # Send email notifications to both complainants
                            email_errors = []
                            
                            try:
                                from backend.services.email_service import message_sender
                                
                                # Get registered case details
                                reg_case = db_queries.get_registered_case_detail(reg_id)
                                if reg_case:
                                    reg_name, reg_mobile, reg_email, reg_age, reg_last_seen, reg_marks, reg_comp_name, reg_city, reg_state, reg_desc = reg_case[0]
                                    
                                    # Send email to registered case complainant
                                    if reg_email:
                                        try:
                                            email_sent = message_sender.send_match_notification(
                                                recipient_email=reg_email,
                                                case_name=reg_name,
                                                matched_person=pub_details.get('submitted_by', 'Unknown'),
                                                location=pub_details.get('location', 'Unknown'),
                                                confidence=f"{confidence:.1f}%",
                                                case_id=reg_id,
                                                registered_case_id=reg_id,
                                                public_case_id=pub_id
                                            )
                                            if email_sent:
                                                st.info(f"📧 Email sent to registered case complainant: {reg_email}")
                                            else:
                                                email_errors.append(f"Failed to send email to {reg_email}")
                                                st.warning(f"⚠️ Failed to send email to registered case complainant: {reg_email}")
                                        except Exception as e:
                                            email_errors.append(f"Error sending email to {reg_email}: {str(e)}")
                                            st.warning(f"⚠️ Error sending email to {reg_email}: {str(e)}")
                                    else:
                                        st.info(f"ℹ️ No email address for registered case complainant")
                                
                                # Get public submission details
                                pub_case = db_queries.get_public_case_detail(pub_id)
                                if pub_case:
                                    pub_name, pub_location, pub_submitted_by, pub_mobile, pub_email, pub_marks, pub_city, pub_state = pub_case[0]
                                    
                                    # Send email to public submission reporter
                                    if pub_email:
                                        try:
                                            email_sent = message_sender.send_match_notification(
                                                recipient_email=pub_email,
                                                case_name=pub_location,
                                                matched_person=reg_details.get('name', 'Unknown'),
                                                location=pub_location,
                                                confidence=f"{confidence:.1f}%",
                                                case_id=pub_id,
                                                registered_case_id=reg_id,
                                                public_case_id=pub_id
                                            )
                                            if email_sent:
                                                st.info(f"📧 Email sent to public submission reporter: {pub_email}")
                                            else:
                                                email_errors.append(f"Failed to send email to {pub_email}")
                                                st.warning(f"⚠️ Failed to send email to public submission reporter: {pub_email}")
                                        except Exception as e:
                                            email_errors.append(f"Error sending email to {pub_email}: {str(e)}")
                                            st.warning(f"⚠️ Error sending email to {pub_email}: {str(e)}")
                                    else:
                                        st.info(f"ℹ️ No email address for public submission reporter")
                                
                                if email_errors:
                                    st.warning(f"⚠️ Some emails could not be sent: {', '.join(email_errors)}")
                                else:
                                    st.success("✅ All emails sent successfully!")
                            
                            except Exception as e:
                                st.warning(f"⚠️ Email notification error: {str(e)}")
                                st.info("✅ Match confirmed, but email notifications could not be sent.")
                                import traceback
                                st.code(traceback.format_exc())
                            
                            st.rerun()
                        except Exception as e:
                            st.error(f"❌ Error confirming match: {str(e)}")
                            import traceback
                            st.code(traceback.format_exc())
                    
                    st.markdown("---")
        else:
                st.error(f"❌ Matching failed: {matched_result.get('message', 'Unknown error')}")
                
                # Show helpful information
                st.markdown("---")
                st.markdown("### 💡 Troubleshooting")
                
                error_msg = matched_result.get('message', '')
                
                if "No registered cases" in error_msg:
                    st.info("📝 **Action Required:** Register at least one missing person case with a clear face photo.")
                elif "No public submissions" in error_msg:
                    st.info("📝 **Information:** Waiting for public submissions. The matching feature will work once public users submit sightings.")
                elif "No valid face mesh" in error_msg:
                    st.info("📷 **Action Required:** Ensure all uploaded photos have clear, visible faces. The system uses face detection to extract facial landmarks.")


def render_send_message():
    """Render send message form"""
    st.markdown("---")
    st.subheader("📧 Send Message to Complainants")
    st.info("Send email notifications to complainants about case updates or send manual messages with attachments.")
    
    from backend.services.email_service import message_sender
    
    # Message mode selection
    message_mode = st.radio(
        "Select Message Mode:",
        ["📋 Case-Based Message", "✉️ Manual Message"],
        horizontal=True
    )
    
    if message_mode == "📋 Case-Based Message":
        # Original case-based message functionality
        render_case_based_message()
    else:
        # New manual message functionality
        render_manual_message()


def render_case_based_message():
    """Render case-based message form"""
    from backend.services.email_service import message_sender
    
    # Get all registered cases
    try:
        all_cases = db_queries.get_all_registered_cases()
        
        if not all_cases:
            st.warning("📭 No registered cases available to send messages.")
            return
        
        # Create case selection options
        case_options = {}
        for case in all_cases:
            status_emoji = "🟢" if case.status == "F" else "🔴"
            case_label = f"{status_emoji} {case.name} - {case.city or 'Unknown'} ({case.complainant_name})"
            case_options[case_label] = case
        
        st.markdown("### 📋 Select Case")
        selected_case_label = st.selectbox(
            "Choose a case to send message about:",
            options=list(case_options.keys())
        )
        
        if selected_case_label:
            selected_case = case_options[selected_case_label]
            
            # Display case details
            with st.expander("📄 Case Details", expanded=True):
                col1, col2 = st.columns(2)
                
                with col1:
                    try:
                        st.image(f"./resources/{selected_case.id}.jpg", width=200)
                    except:
                        st.warning("📷 No image")
                
                with col2:
                    st.write(f"**Name:** {selected_case.name}")
                    st.write(f"**Age:** {selected_case.age or 'Unknown'}")
                    st.write(f"**Last Seen:** {selected_case.last_seen}")
                    st.write(f"**City:** {selected_case.city or 'Not provided'}")
                    st.write(f"**Complainant:** {selected_case.complainant_name}")
                    st.write(f"**Mobile:** {selected_case.complainant_mobile}")
                    st.write(f"**Email:** {selected_case.complainant_email or 'Not provided'}")
                    status_text = "FOUND" if selected_case.status == "F" else "NOT FOUND"
                    st.write(f"**Status:** {status_text}")
            
            st.markdown("---")
            st.markdown("### ✉️ Compose Message")
            
            # Message form
            message_type = st.selectbox(
                "Message Type",
                ["Case Update", "Match Found", "Custom Message"]
            )
            
            if message_type == "Case Update":
                subject = f"Case Update - {selected_case.name}"
                message_body = st.text_area(
                    "Update Message",
                    value=f"We have an update regarding the missing person case for {selected_case.name}.\n\nPlease contact us for more details.",
                    height=200
                )
            
            elif message_type == "Match Found":
                subject = f"MATCH FOUND - {selected_case.name}"
                
                # Get match details if available
                match_info = ""
                if selected_case.matched_with:
                    try:
                        pub_details = db_queries.get_public_case_detail(selected_case.matched_with)
                        if pub_details:
                            pub_detail = pub_details[0]
                            match_info = f"\n\nMatch Details:\n- Location: {pub_detail[1]}\n- Reported By: {pub_detail[2]}\n- Contact: {pub_detail[3]}"
                    except:
                        pass
                
                message_body = st.text_area(
                    "Match Notification",
                    value=f"Great news! We have found a potential match for {selected_case.name}.{match_info}\n\nPlease log in to the system to review the match details and verify the identification.",
                    height=200
                )
            
            else:  # Custom Message
                subject = st.text_input("Subject", value=f"Regarding Case - {selected_case.name}")
                message_body = st.text_area(
                    "Message",
                    value="",
                    height=200,
                    placeholder="Enter your custom message here..."
                )
            
            # Recipient email
            recipient_email = st.text_input(
                "Recipient Email",
                value=selected_case.complainant_email or "",
                help="Email address of the complainant"
            )
            
            # Send button
            col1, col2 = st.columns([1, 3])
            with col1:
                send_button = st.button("📤 Send Email", type="primary", key="send_case_email")
            
            if send_button:
                if not recipient_email or "@" not in recipient_email:
                    st.error("❌ Please provide a valid recipient email address.")
                elif not message_body.strip():
                    st.error("❌ Message body cannot be empty.")
                else:
                    with st.spinner("📧 Sending email..."):
                        try:
                            # Send email using message sender
                            success = message_sender.send_system_notification(
                                recipient_email=recipient_email,
                                title=subject,
                                message=message_body
                            )
                            
                            if success:
                                st.success(f"✅ Email sent successfully to {recipient_email}!")
                                st.balloons()
                                
                                try:
                                    log_user_action(
                                        f"Email sent to {recipient_email}",
                                        st.session_state.get("username")
                                    )
                                except:
                                    pass
                            else:
                                st.error("❌ Failed to send email. Please check email configuration.")
                                st.info("💡 Make sure email settings are configured in email_system/email_config.py")
                        
                        except Exception as e:
                            st.error(f"❌ Error sending email: {str(e)}")
                            st.info("💡 Check email configuration and internet connection.")
    
    except Exception as e:
        st.error(f"❌ Error loading cases: {str(e)}")


def render_manual_message():
    """Render manual message form with attachment support"""
    from backend.services.email_service import message_sender
    
    st.markdown("### ✉️ Manual Message")
    st.info("Send a custom email to any recipient with optional file attachment.")
    
    # Recipient information
    st.markdown("#### 📬 Recipient")
    recipient_email = st.text_input(
        "Recipient Email Address *",
        placeholder="example@email.com",
        help="Enter the email address of the recipient"
    )
    
    st.markdown("---")
    
    # Message content
    st.markdown("#### 📝 Message Content")
    subject = st.text_input(
        "Subject *",
        placeholder="Enter email subject",
        help="Subject line of the email"
    )
    
    message_body = st.text_area(
        "Message *",
        height=250,
        placeholder="Enter your message here...\n\nYou can include:\n- Case updates\n- Instructions\n- Important information\n- Contact details",
        help="Main content of the email"
    )
    
    st.markdown("---")
    
    # Attachment
    st.markdown("#### 📎 Attachment (Optional)")
    attachment_file = st.file_uploader(
        "Upload File",
        type=["pdf", "jpg", "jpeg", "png", "doc", "docx", "txt", "csv", "xlsx"],
        help="Optional: Attach a file to the email (Max 10MB)"
    )
    
    if attachment_file:
        file_size_mb = len(attachment_file.getvalue()) / (1024 * 1024)
        st.info(f"📎 **Attached:** {attachment_file.name} ({file_size_mb:.2f} MB)")
        
        if file_size_mb > 10:
            st.warning("⚠️ File size exceeds 10MB. Consider using a smaller file.")
    
    st.markdown("---")
    
    # Preview
    with st.expander("👁️ Preview Email", expanded=False):
        st.markdown("**To:** " + (recipient_email or "_[Not specified]_"))
        st.markdown("**Subject:** " + (subject or "_[Not specified]_"))
        st.markdown("**Message:**")
        st.text_area(
            "Message Preview",
            value=message_body or "_[Empty message]_",
            height=150,
            disabled=True,
            key="preview",
            label_visibility="collapsed"
        )
        if attachment_file:
            st.markdown(f"**Attachment:** 📎 {attachment_file.name}")
    
    st.markdown("---")
    
    # Send button
    col1, col2, col3 = st.columns([1, 1, 2])
    
    with col1:
        send_button = st.button("📤 Send Email", type="primary", key="send_manual_email")
    
    with col2:
        clear_button = st.button("🗑️ Clear Form", key="clear_manual_form")
    
    if clear_button:
        st.rerun()
    
    if send_button:
        # Validation
        if not recipient_email or "@" not in recipient_email:
            st.error("❌ Please provide a valid recipient email address.")
        elif not subject or not subject.strip():
            st.error("❌ Please provide an email subject.")
        elif not message_body or not message_body.strip():
            st.error("❌ Please provide a message body.")
        else:
            with st.spinner("📧 Sending email..."):
                try:
                    # Save attachment temporarily if provided
                    attachment_path = None
                    attachment_name = None
                    
                    if attachment_file:
                        import tempfile
                        # Create temporary file
                        temp_dir = tempfile.gettempdir()
                        attachment_path = os.path.join(temp_dir, attachment_file.name)
                        attachment_name = attachment_file.name
                        
                        # Write file to temp location
                        with open(attachment_path, "wb") as f:
                            f.write(attachment_file.getvalue())
                        
                        st.info(f"📎 Attaching file: {attachment_name}")
                    
                    # Send email using message sender
                    success = message_sender.send_email_with_attachment(
                        recipient_email=recipient_email,
                        subject=subject,
                        message=message_body,
                        attachment_path=attachment_path,
                        attachment_name=attachment_name
                    )
                    
                    # Clean up temporary file
                    if attachment_path and os.path.exists(attachment_path):
                        try:
                            os.remove(attachment_path)
                        except:
                            pass
                    
                    if success:
                        st.success(f"✅ Email sent successfully to {recipient_email}!")
                        if attachment_file:
                            st.success(f"📎 Attachment '{attachment_name}' included successfully!")
                        st.balloons()
                        
                        try:
                            log_user_action(
                                f"Manual email sent to {recipient_email}" + 
                                (f" with attachment {attachment_name}" if attachment_file else ""),
                                st.session_state.get("username")
                            )
                        except:
                            pass
                        
                        st.info("🔄 Form will remain for sending another message. Click 'Clear Form' to reset.")
                    else:
                        st.error("❌ Failed to send email. Please check email configuration.")
                        st.info("💡 Make sure email settings are configured in email_system/email_config.py")
                
                except Exception as e:
                    st.error(f"❌ Error sending email: {str(e)}")
                    st.info("💡 Check email configuration and internet connection.")
                    import traceback
                    with st.expander("🔍 Error Details"):
                        st.code(traceback.format_exc())


def render_map():
    """Render map visualization"""
    st.markdown("---")
    st.subheader("🗺️ Case Location Map")
    st.info("Interactive map showing locations of both admin registered cases and public submissions.")
    
    try:
        import folium
        from streamlit_folium import st_folium
    except ImportError:
        st.error("❌ Map visualization requires 'folium' and 'streamlit-folium' packages.")
        st.code("pip install folium streamlit-folium", language="bash")
        return
    
    try:
        # Get all registered cases and public submissions
        all_registered = db_queries.get_all_registered_cases()
        all_public = db_queries.fetch_public_cases(train_data=False, status="All")
        
        if not all_registered and not all_public:
            st.warning("📭 No cases to display on map.")
            return
        
        # Filter options
        col1, col2, col3 = st.columns(3)
        
        with col1:
            status_filter = st.selectbox(
                "Filter by Status",
                ["All", "Not Found Only", "Found Only"]
            )
        
        with col2:
            case_type_filter = st.selectbox(
                "Filter by Type",
                ["All", "Admin Cases Only", "Public Submissions Only"]
            )
        
        with col3:
            city_filter = st.text_input("Filter by City", "")
        
        # Combine cases
        all_cases = []
        
        # Add registered cases
        for case in all_registered:
            all_cases.append({
                'type': 'registered',
                'name': case.name,
                'city': case.city,
                'status': case.status,
                'age': case.age,
                'last_seen': case.last_seen,
                'complainant_name': case.complainant_name,
                'complainant_mobile': case.complainant_mobile,
                'marker_color': 'blue'
            })
        
        # Add public submissions
        for sub in all_public:
            all_cases.append({
                'type': 'public',
                'name': sub.name if (hasattr(sub, 'name') and sub.name) else 'Unknown',
                'city': sub.city if hasattr(sub, 'city') else None,
                'status': sub.status,
                'age': 'Unknown',
                'last_seen': sub.location,
                'complainant_name': sub.submitted_by or 'Anonymous',
                'complainant_mobile': sub.mobile,
                'marker_color': 'orange'
            })
        
        # Apply filters
        filtered_cases = all_cases
        
        if status_filter == "Not Found Only":
            filtered_cases = [c for c in filtered_cases if c['status'] == "NF"]
        elif status_filter == "Found Only":
            filtered_cases = [c for c in filtered_cases if c['status'] == "F"]
        
        if case_type_filter == "Admin Cases Only":
            filtered_cases = [c for c in filtered_cases if c['type'] == 'registered']
        elif case_type_filter == "Public Submissions Only":
            filtered_cases = [c for c in filtered_cases if c['type'] == 'public']
        
        if city_filter:
            filtered_cases = [
                c for c in filtered_cases
                if c['city'] and city_filter.lower() in c['city'].lower()
            ]
        
        st.markdown("---")
        
        # Create map centered on India (approximate center)
        m = folium.Map(
            location=[20.5937, 78.9629],  # Center of India
            zoom_start=5,
            tiles="OpenStreetMap"
        )
        
        # City coordinates (approximate - in real implementation, use geocoding API)
        city_coords = {
            "mumbai": [19.0760, 72.8777],
            "delhi": [28.7041, 77.1025],
            "bangalore": [12.9716, 77.5946],
            "hyderabad": [17.3850, 78.4867],
            "chennai": [13.0827, 80.2707],
            "kolkata": [22.5726, 88.3639],
            "pune": [18.5204, 73.8567],
            "ahmedabad": [23.0225, 72.5714],
            "jaipur": [26.9124, 75.7873],
            "lucknow": [26.8467, 80.9462],
            "chh. sambhajinagar": [19.8762, 75.3433],
            "chhatrapati sambhajinagar": [19.8762, 75.3433],
            "aurangabad": [19.8762, 75.3433],
            "nagpur": [21.1458, 79.0882],
            "nashik": [19.9975, 73.7898],
            "thane": [19.2183, 72.9781],
            "beed": [19.2403, 75.7597],
            "parbhani": [19.2683, 76.7597],
            "latur": [18.4088, 76.5158],
            "nanded": [19.1383, 77.3267],
            "washim": [20.1033, 76.8197],
            "akola": [20.7136, 77.0064],
            "amravati": [20.8530, 77.7539],
            "yavatmal": [20.4183, 78.1192],
            "buldhana": [20.5497, 76.1744],
            "jalna": [19.8450, 75.8783],
            "solapur": [17.6599, 75.9064],
            "sangli": [16.8633, 74.5681],
            "satara": [17.6726, 73.9190],
            "kolhapur": [16.7050, 74.2433],
            "ratnagiri": [16.9891, 73.3167],
            "sindhudurg": [16.0158, 73.5244],
            "raigad": [18.5244, 73.2600],
            "raigarh": [21.8974, 83.4086],
        }
        
        # Add markers for each case
        markers_added = 0
        for case in filtered_cases:
            if not case['city']:
                continue
            
            # Get coordinates for city
            city_key = case['city'].lower().strip()
            coords = city_coords.get(city_key)
            
            if not coords:
                # Try partial match
                for key, coord in city_coords.items():
                    if key in city_key or city_key in key:
                        coords = coord
                        break
            
            # If still no match, try removing common suffixes
            if not coords:
                city_variations = [
                    city_key.replace(" city", ""),
                    city_key.replace(" district", ""),
                    city_key.replace(" taluka", ""),
                    city_key.replace(" tehsil", ""),
                ]
                for variation in city_variations:
                    if variation in city_coords:
                        coords = city_coords[variation]
                        break
                    # Try partial match on variations
                    for key, coord in city_coords.items():
                        if key in variation or variation in key:
                            coords = coord
                            break
                    if coords:
                        break
            
            if coords:
                # Determine marker color based on status and type
                if case['status'] == "F":
                    color = "green"
                    icon = "ok-sign"
                else:
                    color = case['marker_color']  # blue for admin, orange for public
                    icon = "exclamation-sign"
                
                # Create popup content
                status_text = "FOUND" if case['status'] == "F" else "NOT FOUND"
                case_type_text = "🏢 Admin Case" if case['type'] == 'registered' else "👥 Public Submission"
                
                popup_html = f"""
                <div style="width: 220px;">
                    <h4>{case['name']}</h4>
                    <p><b>Type:</b> {case_type_text}</p>
                    <p><b>Status:</b> {status_text}</p>
                    <p><b>Age:</b> {case['age']}</p>
                    <p><b>City:</b> {case['city']}</p>
                    <p><b>Last Seen/Location:</b> {case['last_seen']}</p>
                    <p><b>Reported By:</b> {case['complainant_name']}</p>
                    <p><b>Mobile:</b> {case['complainant_mobile']}</p>
                </div>
                """
                
                # Add marker
                folium.Marker(
                    location=coords,
                    popup=folium.Popup(popup_html, max_width=250),
                    tooltip=f"{case['name']} - {case['city']}",
                    icon=folium.Icon(color=color, icon=icon, prefix='glyphicon')
                ).add_to(m)
                
                markers_added += 1
        
        # Display statistics
        st.write(f"**Total Cases:** {len(filtered_cases)} | **Markers on Map:** {markers_added}")
        
        if markers_added == 0:
            st.warning("⚠️ No cases could be mapped. City coordinates not available for filtered cases.")
            st.info("💡 Supported cities: Mumbai, Delhi, Bangalore, Hyderabad, Chennai, Kolkata, Pune, Ahmedabad, Jaipur, Lucknow, Chh. Sambhajinagar, Nagpur, Nashik, Thane")
        
        # Display map
        st.markdown("### 🗺️ Interactive Map")
        st.caption("🔴 Red markers = Not Found | 🟢 Green markers = Found | 🔵 Blue markers = Admin (Not Found) | 🟠 Orange markers = Public (Not Found) | Click markers for details")
        
        st_folium(m, width=1200, height=600)
        
            
    except Exception as e:
        st.error(f"❌ Error rendering map: {str(e)}")
        import traceback
        st.code(traceback.format_exc())


def render_email_configuration():
    """Render email configuration page"""
    st.markdown("---")
    st.subheader("⚙️ Email Configuration")
    st.info("Configure email settings for sending notifications and messages.")
    
    from email_system.email_config import EmailConfig
    
    # Load current configuration
    email_config = EmailConfig()
    current_config = email_config.get_config()
    is_configured = email_config.is_configured()
    
    # Display current status
    st.markdown("### 📊 Current Status")
    
    col1, col2 = st.columns(2)
    
    with col1:
        if is_configured:
            st.success("✅ Email is configured")
        else:
            st.error("❌ Email is not configured")
    
    with col2:
        if st.button("🧪 Test Email Configuration"):
            with st.spinner("Testing email configuration..."):
                from backend.services.email_service import message_sender
                test_result = message_sender.test_email_config()
                
                if test_result.get('configured'):
                    st.success("✅ Email configuration is valid!")
                    st.write(f"**SMTP Host:** {test_result.get('smtp_host')}")
                    st.write(f"**SMTP Port:** {test_result.get('smtp_port')}")
                    st.write(f"**SMTP User:** {test_result.get('smtp_user')}")
                else:
                    st.error("❌ Email configuration test failed")
                    if test_result.get('error'):
                        st.error(f"**Error:** {test_result.get('error')}")
    
    st.markdown("---")
    
    # Configuration tabs
    tab1, tab2, tab3 = st.tabs(["📝 Configure Email", "📖 Instructions", "🔍 Troubleshooting"])
    
    with tab1:
        st.markdown("### 📝 Email Settings")
        st.info("💡 For Gmail, you need to use an **App Password**, not your regular password. See Instructions tab for details.")
        
        # Configuration form
        with st.form("email_config_form"):
            st.markdown("#### SMTP Server Settings")
            
            smtp_provider = st.selectbox(
                "Email Provider",
                ["Gmail", "Outlook", "Yahoo", "Custom"],
                help="Select your email provider or choose Custom for other providers"
            )
            
            # Pre-fill based on provider
            if smtp_provider == "Gmail":
                default_host = "smtp.gmail.com"
                default_port = 587
            elif smtp_provider == "Outlook":
                default_host = "smtp-mail.outlook.com"
                default_port = 587
            elif smtp_provider == "Yahoo":
                default_host = "smtp.mail.yahoo.com"
                default_port = 587
            else:
                default_host = current_config.get('smtp_host', '')
                default_port = current_config.get('smtp_port', 587)
            
            smtp_host = st.text_input(
                "SMTP Host *",
                value=default_host,
                help="SMTP server address (e.g., smtp.gmail.com)"
            )
            
            smtp_port = st.number_input(
                "SMTP Port *",
                value=default_port,
                min_value=1,
                max_value=65535,
                help="SMTP port (usually 587 for TLS or 465 for SSL)"
            )
            
            st.markdown("#### Email Account")
            
            smtp_user = st.text_input(
                "Email Address *",
                value=current_config.get('smtp_user', ''),
                placeholder="your-email@gmail.com",
                help="Your email address"
            )
            
            smtp_password = st.text_input(
                "Password / App Password *",
                type="password",
                help="For Gmail: Use App Password (see Instructions tab)"
            )
            
            notify_email = st.text_input(
                "Notification Email (Optional)",
                value=current_config.get('notify_email', ''),
                placeholder="admin@example.com",
                help="Optional: Email to receive system notifications"
            )
            
            st.markdown("---")
            
            col1, col2 = st.columns(2)
            
            with col1:
                save_button = st.form_submit_button("💾 Save Configuration", type="primary")
            
            with col2:
                save_to_env = st.checkbox("Save to .env file", value=False, help="Save to .env file instead of email_settings.txt")
            
            if save_button:
                if not smtp_host or not smtp_user or not smtp_password:
                    st.error("❌ Please fill in all required fields (marked with *)")
                else:
                    try:
                        if save_to_env:
                            # Save to .env file
                            env_content = f"""# Email Configuration
SMTP_HOST={smtp_host}
SMTP_PORT={smtp_port}
SMTP_USER={smtp_user}
SMTP_PASSWORD={smtp_password}
NOTIFY_EMAIL={notify_email}
"""
                            with open(".env", "w") as f:
                                f.write(env_content)
                            st.success("✅ Configuration saved to .env file!")
                        else:
                            # Save to email_settings.txt
                            config_content = f"""# Email Configuration for Missing Person Identification System
# Edit these settings with your email provider details

# SMTP Server Settings
SMTP_HOST={smtp_host}
SMTP_PORT={smtp_port}
SMTP_USER={smtp_user}
SMTP_PASSWORD={smtp_password}

# Default notification email (optional)
NOTIFY_EMAIL={notify_email}

# Common SMTP settings:
# Gmail: smtp.gmail.com:587
# Outlook: smtp-mail.outlook.com:587
# Yahoo: smtp.mail.yahoo.com:587
"""
                            with open("email_settings.txt", "w") as f:
                                f.write(config_content)
                            st.success("✅ Configuration saved to email_settings.txt!")
                        
                        st.balloons()
                        st.info("🔄 Please restart the application for changes to take effect.")
                        
                        # Log action
                        try:
                            log_user_action(
                                "Email configuration updated",
                                st.session_state.get("username")
                            )
                        except:
                            pass
                    
                    except Exception as e:
                        st.error(f"❌ Error saving configuration: {str(e)}")
    
    with tab2:
        st.markdown("### 📖 How to Configure Email")
        
        st.markdown("#### For Gmail Users (Recommended)")
        st.markdown("""
1. **Enable 2-Step Verification**
   - Go to your Google Account settings
   - Navigate to Security → 2-Step Verification
   - Enable it if not already enabled

2. **Generate App Password**
   - Go to: https://myaccount.google.com/apppasswords
   - Select "Mail" and "Other (Custom name)"
   - Enter "Missing Person System" as the name
   - Click "Generate"
   - Copy the 16-character password (remove spaces)

3. **Configure in System**
   - Email Provider: Gmail
   - SMTP Host: smtp.gmail.com
   - SMTP Port: 587
   - Email Address: your-email@gmail.com
   - Password: [Paste the 16-character App Password]

4. **Save and Test**
   - Click "Save Configuration"
   - Click "Test Email Configuration"
   - If successful, you're ready to send emails!
""")
        
        st.markdown("---")
        
        st.markdown("#### For Outlook/Hotmail Users")
        st.markdown("""
1. **SMTP Settings**
   - SMTP Host: smtp-mail.outlook.com
   - SMTP Port: 587
   - Email: your-email@outlook.com
   - Password: Your Outlook password

2. **Note**: Outlook may require additional security settings
""")
        
        st.markdown("---")
        
        st.markdown("#### For Yahoo Users")
        st.markdown("""
1. **Generate App Password**
   - Go to Yahoo Account Security
   - Generate an app password for "Mail"

2. **SMTP Settings**
   - SMTP Host: smtp.mail.yahoo.com
   - SMTP Port: 587
   - Email: your-email@yahoo.com
   - Password: [App Password]
""")
        
        st.markdown("---")
        
        st.markdown("#### Security Notes")
        st.warning("""
⚠️ **Important Security Information:**
- Never share your App Password with anyone
- The password is stored locally in email_settings.txt or .env
- Keep these files secure and never commit them to version control
- Add email_settings.txt and .env to .gitignore
""")
    
    with tab3:
        st.markdown("### 🔍 Troubleshooting")
        
        st.markdown("#### Common Issues")
        
        with st.expander("❌ Authentication Failed (535 Error)"):
            st.markdown("""
**Problem**: Username and Password not accepted

**Solutions**:
1. For Gmail: Make sure you're using an **App Password**, not your regular password
2. Check that 2-Step Verification is enabled on your Google Account
3. Verify the email address is correct
4. Regenerate the App Password and try again
5. Check for extra spaces in the password
""")
        
        with st.expander("❌ Connection Failed"):
            st.markdown("""
**Problem**: Cannot connect to SMTP server

**Solutions**:
1. Check your internet connection
2. Verify SMTP host and port are correct
3. Check if your firewall is blocking port 587
4. Try port 465 with SSL instead of 587 with TLS
5. Ensure your email provider allows SMTP access
""")
        
        with st.expander("❌ Email Not Configured"):
            st.markdown("""
**Problem**: System says email is not configured

**Solutions**:
1. Make sure you saved the configuration
2. Restart the application after saving
3. Check that email_settings.txt or .env file exists
4. Verify all required fields are filled
5. Check file permissions
""")
        
        with st.expander("❌ Emails Not Being Received"):
            st.markdown("""
**Problem**: Emails sent successfully but not received

**Solutions**:
1. Check recipient's spam/junk folder
2. Verify recipient email address is correct
3. Check your email provider's sending limits
4. Ensure your email account is not suspended
5. Try sending a test email to yourself first
""")
        
        st.markdown("---")
        
        st.markdown("#### Test Configuration")
        st.info("""
💡 **Testing Tips**:
1. Save your configuration first
2. Click "Test Email Configuration" button
3. If test passes, try sending a test email to yourself
4. Check both inbox and spam folder
5. If issues persist, review error messages carefully
""")
        
        st.markdown("---")
        
        st.markdown("#### Need More Help?")
        st.markdown("""
- Check application logs in `logs/` folder
- Review error messages in the Send Message page
- Ensure email provider allows third-party app access
- Contact your email provider's support if needed
""")
    
    st.markdown("---")
    
    # Quick actions
    st.markdown("### ⚡ Quick Actions")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        if st.button("📧 Send Test Email"):
            test_email = st.text_input("Enter test email address:", key="test_email_input")
            if test_email and "@" in test_email:
                with st.spinner("Sending test email..."):
                    from backend.services.email_service import message_sender
                    success = message_sender.send_system_notification(
                        recipient_email=test_email,
                        title="Test Email from Missing Person System",
                        message="This is a test email to verify your email configuration is working correctly.\n\nIf you received this email, your email system is configured properly!"
                    )
                    
                    if success:
                        st.success(f"✅ Test email sent to {test_email}!")
                        st.info("Check your inbox (and spam folder)")
                    else:
                        st.error("❌ Failed to send test email")
    
    with col2:
        if st.button("📄 View Current Config"):
            st.json({
                "smtp_host": current_config.get('smtp_host', 'Not set'),
                "smtp_port": current_config.get('smtp_port', 'Not set'),
                "smtp_user": current_config.get('smtp_user', 'Not set'),
                "has_password": bool(current_config.get('smtp_password')),
                "notify_email": current_config.get('notify_email', 'Not set')
            })
    
    with col3:
        if st.button("🗑️ Clear Configuration"):
            if st.checkbox("Confirm deletion", key="confirm_delete_config"):
                try:
                    if os.path.exists("email_settings.txt"):
                        os.remove("email_settings.txt")
                        st.success("✅ email_settings.txt deleted")
                    if os.path.exists(".env"):
                        os.remove(".env")
                        st.success("✅ .env deleted")
                    st.info("🔄 Please restart the application")
                except Exception as e:
                    st.error(f"❌ Error: {str(e)}")



def render_user_management():
    """Render user management page - View, search, and manage registered users"""
    st.markdown("---")
    st.markdown("""
    <div style='background: linear-gradient(135deg, #9b59b6 0%, #8e44ad 100%); 
                padding: 1.5rem; border-radius: 8px; margin-bottom: 1.5rem;
                box-shadow: 0 2px 8px rgba(0,0,0,0.1);'>
        <h2 style='color: white; margin: 0; font-size: 1.8rem;'>👥 User Management</h2>
        <p style='color: rgba(255,255,255,0.9); margin: 0.5rem 0 0 0;'>
            View, search, and manage registered public users
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    try:
        from backend.services.auth_service import UserAuth
        
        # Initialize auth service
        auth = UserAuth()
        
        # Get all users
        all_users = auth.get_all_users()
        
        # Get user statistics
        try:
            stats = auth.get_user_stats()
        except:
            # Calculate stats manually if method doesn't exist
            stats = {
                "total_users": len(all_users),
                "verified_users": sum(1 for u in all_users if u.get("verified") == "Yes"),
                "unverified_users": sum(1 for u in all_users if u.get("verified") == "No")
            }
        
        # Display statistics
        st.markdown("### 📊 User Statistics")
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.markdown(f"""
            <div style='background: linear-gradient(135deg, #3498db 0%, #2980b9 100%); 
                        padding: 1.5rem; border-radius: 8px; text-align: center;
                        box-shadow: 0 2px 8px rgba(0,0,0,0.1);'>
                <p style='color: rgba(255,255,255,0.8); margin: 0; font-size: 0.9rem;'>Total Users</p>
                <p style='color: white; margin: 0.5rem 0 0 0; font-size: 2rem; font-weight: bold;'>
                    {stats.get('total_users', 0)}
                </p>
            </div>
            """, unsafe_allow_html=True)
        
        with col2:
            st.markdown(f"""
            <div style='background: linear-gradient(135deg, #27ae60 0%, #229954 100%); 
                        padding: 1.5rem; border-radius: 8px; text-align: center;
                        box-shadow: 0 2px 8px rgba(0,0,0,0.1);'>
                <p style='color: rgba(255,255,255,0.8); margin: 0; font-size: 0.9rem;'>Verified</p>
                <p style='color: white; margin: 0.5rem 0 0 0; font-size: 2rem; font-weight: bold;'>
                    {stats.get('verified_users', 0)}
                </p>
            </div>
            """, unsafe_allow_html=True)
        
        with col3:
            st.markdown(f"""
            <div style='background: linear-gradient(135deg, #e74c3c 0%, #c0392b 100%); 
                        padding: 1.5rem; border-radius: 8px; text-align: center;
                        box-shadow: 0 2px 8px rgba(0,0,0,0.1);'>
                <p style='color: rgba(255,255,255,0.8); margin: 0; font-size: 0.9rem;'>Unverified</p>
                <p style='color: white; margin: 0.5rem 0 0 0; font-size: 2rem; font-weight: bold;'>
                    {stats.get('unverified_users', 0)}
                </p>
            </div>
            """, unsafe_allow_html=True)
        
        st.markdown("---")
        
        # Search and filter
        st.markdown("### 🔍 Search & Filter")
        col1, col2 = st.columns(2)
        
        with col1:
            search_term = st.text_input("Search by email or name", "", key="user_search")
        
        with col2:
            filter_verified = st.selectbox(
                "Filter by verification status",
                ["All", "Verified", "Unverified"],
                key="user_filter"
            )
        
        # Apply filters
        filtered_users = all_users
        
        # Search filter
        if search_term:
            search_lower = search_term.lower()
            filtered_users = [
                u for u in filtered_users 
                if search_lower in u.get("email", "").lower() 
                or search_lower in u.get("name", "").lower()
            ]
        
        # Verification filter
        if filter_verified == "Verified":
            filtered_users = [u for u in filtered_users if u.get("verified") == "Yes"]
        elif filter_verified == "Unverified":
            filtered_users = [u for u in filtered_users if u.get("verified") == "No"]
        
        st.markdown("---")
        
        # Display users
        st.markdown(f"### 👤 Registered Users ({len(filtered_users)})")
        
        if not filtered_users:
            st.info("No users found matching your search criteria.")
        else:
            # Display users in expandable cards
            for idx, user in enumerate(filtered_users):
                user_id = user.get("id")
                email = user.get("email", "N/A")
                name = user.get("name", "N/A")
                phone = user.get("phone", "N/A")
                verified = user.get("verified", "No")
                created_at = user.get("created_at", "N/A")
                last_login = user.get("last_login", "Never")
                
                # Status badge
                if verified == "Yes":
                    status_badge = "✅ Verified"
                    status_color = "#27ae60"
                else:
                    status_badge = "⚠️ Unverified"
                    status_color = "#e74c3c"
                
                # Create expander title
                title = f"{status_badge} | {name} ({email})"
                
                with st.expander(title, expanded=False):
                    col_info, col_actions = st.columns([2, 1])
                    
                    with col_info:
                        st.markdown("**📧 Contact Information**")
                        st.write(f"**Email:** {email}")
                        st.write(f"**Name:** {name}")
                        st.write(f"**Phone:** {phone}")
                        
                        st.markdown("**📊 Account Status**")
                        st.markdown(f"**Status:** <span style='color: {status_color}; font-weight: bold;'>{status_badge}</span>", unsafe_allow_html=True)
                        st.write(f"**Created:** {created_at}")
                        st.write(f"**Last Login:** {last_login}")
                        
                        # Get user's submissions count
                        try:
                            user_submissions = db_queries.fetch_public_cases(train_data=False, status="All")
                            user_submission_count = sum(1 for s in user_submissions if s.submitted_by == email)
                            st.write(f"**Submissions:** {user_submission_count} cases")
                        except:
                            pass
                    
                    with col_actions:
                        st.markdown("**⚙️ Actions:**")
                        
                        # View submissions button
                        if st.button("📋 View Submissions", key=f"view_sub_{user_id}_{idx}"):
                            st.session_state[f"viewing_submissions_{user_id}"] = True
                            st.rerun()
                        
                        # Delete user button
                        st.markdown("---")
                        st.markdown("**⚠️ Danger Zone:**")
                        
                        if st.button("🗑️ Delete User", key=f"del_user_{user_id}_{idx}", type="secondary"):
                            st.session_state[f"confirm_delete_{user_id}"] = True
                        
                        # Confirmation dialog
                        if st.session_state.get(f"confirm_delete_{user_id}", False):
                            st.warning(f"⚠️ Delete user **{email}**?")
                            st.caption("This will delete the user account and all their submissions permanently.")
                            
                            col_yes, col_no = st.columns(2)
                            
                            with col_yes:
                                if st.button("✅ Yes, Delete", key=f"confirm_yes_{user_id}_{idx}", type="primary"):
                                    try:
                                        # Delete user
                                        success, message = auth.delete_user(user_id)
                                        
                                        if success:
                                            st.success(message)
                                            
                                            try:
                                                log_user_action(f"Deleted user: {email}", st.session_state.get("username"))
                                            except:
                                                pass
                                            
                                            # Clear Streamlit cache to ensure fresh data
                                            st.cache_data.clear()
                                            st.cache_resource.clear()
                                            
                                            # Clear state and refresh
                                            if f"confirm_delete_{user_id}" in st.session_state:
                                                del st.session_state[f"confirm_delete_{user_id}"]
                                            
                                            st.info("✅ User deleted! Please refresh the page or restart Streamlit for changes to take full effect.")
                                            st.rerun()
                                        else:
                                            st.error(message)
                                    except Exception as e:
                                        st.error(f"❌ Error: {str(e)}")
                            
                            with col_no:
                                if st.button("❌ Cancel", key=f"confirm_no_{user_id}_{idx}"):
                                    del st.session_state[f"confirm_delete_{user_id}"]
                                    st.rerun()
                    
                    # View submissions section
                    if st.session_state.get(f"viewing_submissions_{user_id}", False):
                        st.markdown("---")
                        st.markdown(f"### 📋 Submissions by {name}")
                        
                        try:
                            user_submissions = db_queries.fetch_public_cases(train_data=False, status="All")
                            user_cases = [s for s in user_submissions if s.submitted_by == email]
                            
                            if user_cases:
                                for sub_idx, submission in enumerate(user_cases):
                                    sub_id = submission.id
                                    sub_name = submission.name if hasattr(submission, 'name') else 'Unknown'
                                    sub_location = submission.location
                                    sub_status = submission.status
                                    
                                    status_emoji = "✅" if sub_status == "F" else "🔍"
                                    status_text = "Found" if sub_status == "F" else "Not Found"
                                    
                                    st.markdown(f"**{status_emoji} {sub_name}** - {sub_location} ({status_text})")
                                    
                                    sub_col1, sub_col2 = st.columns([3, 1])
                                    
                                    with sub_col1:
                                        st.caption(f"Submitted: {format_time_12h(submission.submitted_on)}")
                                    
                                    with sub_col2:
                                        if st.button("🗑️ Delete", key=f"del_sub_{sub_id}_{sub_idx}"):
                                            try:
                                                db_queries.delete_public_submission(sub_id)
                                                st.success("Submission deleted!")
                                                st.rerun()
                                            except Exception as e:
                                                st.error(f"Error: {str(e)}")
                                    
                                    st.markdown("---")
                            else:
                                st.info("No submissions found for this user.")
                        except Exception as e:
                            st.error(f"Error loading submissions: {str(e)}")
                        
                        if st.button("⬅️ Back to User Details", key=f"back_user_{user_id}"):
                            del st.session_state[f"viewing_submissions_{user_id}"]
                            st.rerun()
        
        st.markdown("---")
        
        # Bulk actions
        st.markdown("### ⚡ Bulk Actions")
        st.warning("⚠️ Use these actions with caution!")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            if st.button("📊 Export User List", type="secondary"):
                import io
                try:
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, Alignment, PatternFill
                except ImportError:
                    st.error("❌ openpyxl library not installed. Please install it using: pip install openpyxl")
                    st.stop()
                
                # Create Excel workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "User List"
                
                # Write project title in first row
                ws['A1'] = 'AI MISSING PERSON IDENTIFICATION SYSTEM - Registered Users & Submissions'
                ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
                ws['A1'].fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
                ws.merge_cells('A1:N1')
                ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
                ws.row_dimensions[1].height = 25
                
                # Write column headers in row 2
                headers = [
                    'Sr.No',
                    'User ID',
                    'Email',
                    'Name',
                    'Phone',
                    'Verified',
                    'Created At',
                    'Last Login',
                    'Total Submissions',
                    'Submission ID',
                    'Person Name',
                    'Location',
                    'Status',
                    'Submitted On'
                ]
                
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=2, column=col_idx)
                    cell.value = header
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Get all public submissions
                try:
                    all_submissions = db_queries.fetch_public_cases(train_data=False, status="All")
                except:
                    all_submissions = []
                
                # Create user submissions map
                user_submissions_map = {}
                for sub in all_submissions:
                    email = sub.email if hasattr(sub, 'email') else sub.submitted_by
                    if email not in user_submissions_map:
                        user_submissions_map[email] = []
                    user_submissions_map[email].append(sub)
                
                # Write data starting from row 3
                sr_no = 1
                row_num = 3
                
                for user in all_users:
                    user_id = user['id']
                    email = user['email']
                    name = user['name']
                    phone = user['phone']
                    verified = user['verified']
                    created_at = user['created_at']
                    last_login = user['last_login']
                    
                    # Get user's submissions
                    user_subs = user_submissions_map.get(email, [])
                    total_subs = len(user_subs)
                    
                    if user_subs:
                        # Write one row per submission
                        for sub in user_subs:
                            status_text = "FOUND" if sub.status == "F" else "NOT FOUND"
                            person_name = sub.name if (hasattr(sub, 'name') and sub.name) else 'Unknown'
                            location = sub.location if hasattr(sub, 'location') else ''
                            submitted_on = sub.submitted_on.strftime('%Y-%m-%d %I:%M %p') if hasattr(sub, 'submitted_on') else ''
                            
                            row_data = [
                                sr_no,
                                user_id,
                                email,
                                name,
                                phone,
                                verified,
                                created_at,
                                last_login,
                                total_subs,
                                sub.id,
                                person_name,
                                location,
                                status_text,
                                submitted_on
                            ]
                            
                            for col_idx, value in enumerate(row_data, 1):
                                ws.cell(row=row_num, column=col_idx).value = value
                            
                            sr_no += 1
                            row_num += 1
                    else:
                        # User with no submissions
                        row_data = [
                            sr_no,
                            user_id,
                            email,
                            name,
                            phone,
                            verified,
                            created_at,
                            last_login,
                            0,
                            '',
                            '',
                            '',
                            '',
                            ''
                        ]
                        
                        for col_idx, value in enumerate(row_data, 1):
                            ws.cell(row=row_num, column=col_idx).value = value
                        
                        sr_no += 1
                        row_num += 1
                
                # Adjust column widths
                column_widths = [8, 10, 30, 20, 15, 10, 20, 20, 15, 36, 20, 25, 12, 20]
                for i, width in enumerate(column_widths, 1):
                    ws.column_dimensions[chr(64 + i)].width = width
                
                # Save to bytes
                excel_buffer = io.BytesIO()
                wb.save(excel_buffer)
                excel_buffer.seek(0)
                
                # Download button
                st.download_button(
                    label="⬇️ Download Excel",
                    data=excel_buffer.getvalue(),
                    file_name=f"users_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        
        with col2:
            if st.button("🔄 Refresh User List", type="secondary"):
                st.cache_data.clear()
                st.cache_resource.clear()
                st.rerun()
        
        with col3:
            if st.button("🧹 Clear Cache", type="secondary"):
                st.cache_data.clear()
                st.cache_resource.clear()
                st.success("✅ Cache cleared! Registration should work now.")
                st.info("💡 If 'Email already registered' still appears, restart Streamlit.")
    
    except ImportError:
        st.error("❌ User authentication system not available")
        st.info("The user management feature requires the authentication service to be properly configured.")
    except Exception as e:
        st.error(f"❌ Error loading user management: {str(e)}")
        st.info("Please check the application logs for more details.")
